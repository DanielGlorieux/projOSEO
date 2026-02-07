"""
Service d'export de données vers Excel
Pour téléchargement et analyse externe
"""
import pandas as pd
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, BarChart

logger = logging.getLogger(__name__)

DATA_PATH = Path(__file__).parent.parent / "data" / "raw"
EXPORT_PATH = Path(__file__).parent.parent / "data" / "exports"


class ExcelExportService:
    """Service d'export de datasets au format Excel"""
    
    def __init__(self):
        # Créer le dossier d'exports s'il n'existe pas
        EXPORT_PATH.mkdir(parents=True, exist_ok=True)
    
    def load_station_data(self, station_id: str, days: Optional[int] = None) -> pd.DataFrame:
        """Charger les données d'une station"""
        csv_path = DATA_PATH / f"{station_id}_historical.csv"
        
        if not csv_path.exists():
            raise FileNotFoundError(f"Données non trouvées pour {station_id}")
        
        df = pd.read_csv(csv_path)
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df = df.sort_values('timestamp')
        
        # Filtrer par période si spécifié
        if days:
            cutoff_date = datetime.now() - timedelta(days=days)
            df = df[df['timestamp'] >= cutoff_date]
        
        return df
    
    def format_excel_worksheet(self, ws, title: str):
        """Formatter une feuille Excel avec styles professionnels"""
        # Titre
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30
        
        # En-têtes de colonnes
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[3]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    def export_station_data(
        self,
        station_id: str,
        station_name: str,
        days: Optional[int] = 7,
        include_analytics: bool = True
    ) -> str:
        """
        Exporter les données d'une station au format Excel avec analyses
        
        Args:
            station_id: ID de la station
            station_name: Nom de la station
            days: Nombre de jours de données (None = toutes)
            include_analytics: Inclure feuilles d'analyse
            
        Returns:
            Chemin du fichier Excel généré
        """
        try:
            # Charger les données
            df = self.load_station_data(station_id, days)
            
            # Vérifier les colonnes nécessaires
            required_cols = ['timestamp', 'energy_consumption_kwh', 'energy_cost_fcfa', 'pump_efficiency']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                logger.warning(f"Colonnes manquantes: {missing_cols}, export basique seulement")
                include_analytics = False
            
            # Nom du fichier
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ONEA_{station_id}_{timestamp}.xlsx"
            filepath = EXPORT_PATH / filename
            
            # Créer le workbook Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # === Feuille 1: Données brutes ===
                df_export = df.copy()
                df_export['timestamp'] = df_export['timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
                df_export.to_excel(writer, sheet_name='Données Brutes', index=False, startrow=2)
                
                ws_raw = writer.sheets['Données Brutes']
                self.format_excel_worksheet(ws_raw, f"Station {station_name} - Données Brutes")
                
                # Ajuster largeur des colonnes
                for column in ws_raw.columns:
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws_raw.column_dimensions[column_letter].width = min(max_length + 2, 30)
                
                if include_analytics:
                    # === Feuille 2: Résumé Statistique ===
                    stats_data = {
                        'Métrique': [
                            'Période analysée',
                            'Nombre de points de données',
                            'Consommation totale (kWh)',
                            'Consommation moyenne (kWh)',
                            'Coût total (FCFA)',
                            'Coût moyen (FCFA)',
                            'Efficacité moyenne (%)',
                            'Nombre d\'anomalies',
                            'Pompes actives (moyenne)',
                            'Facteur de puissance moyen'
                        ],
                        'Valeur': [
                            f"{df['timestamp'].min().strftime('%Y-%m-%d')} à {df['timestamp'].max().strftime('%Y-%m-%d')}",
                            len(df),
                            f"{df['energy_consumption_kwh'].sum():,.2f}",
                            f"{df['energy_consumption_kwh'].mean():,.2f}",
                            f"{df['energy_cost_fcfa'].sum():,.2f}",
                            f"{df['energy_cost_fcfa'].mean():,.2f}",
                            f"{df['pump_efficiency'].mean() * 100:.2f}",
                            df['anomaly'].sum() if 'anomaly' in df.columns else 0,
                            f"{df['pumps_active'].mean():.2f}",
                            f"{df['power_factor'].mean():.3f}" if 'power_factor' in df.columns else "N/A"
                        ]
                    }
                    
                    df_stats = pd.DataFrame(stats_data)
                    df_stats.to_excel(writer, sheet_name='Résumé Statistique', index=False, startrow=2)
                    
                    ws_stats = writer.sheets['Résumé Statistique']
                    self.format_excel_worksheet(ws_stats, f"Station {station_name} - Résumé Statistique")
                    ws_stats.column_dimensions['A'].width = 35
                    ws_stats.column_dimensions['B'].width = 30
                    
                    # === Feuille 3: Analyse par Heure ===
                    df['hour'] = pd.to_datetime(df['timestamp']).dt.hour
                    hourly_analysis = df.groupby('hour').agg({
                        'energy_consumption_kwh': ['mean', 'min', 'max', 'std'],
                        'energy_cost_fcfa': ['mean', 'sum'],
                        'pump_efficiency': 'mean',
                        'pumps_active': 'mean'
                    }).round(2)
                    
                    hourly_analysis.columns = ['_'.join(col).strip() for col in hourly_analysis.columns]
                    hourly_analysis.reset_index(inplace=True)
                    hourly_analysis.to_excel(writer, sheet_name='Analyse Horaire', index=False, startrow=2)
                    
                    ws_hourly = writer.sheets['Analyse Horaire']
                    self.format_excel_worksheet(ws_hourly, f"Station {station_name} - Analyse par Heure")
                    
                    for column in ws_hourly.columns:
                        ws_hourly.column_dimensions[get_column_letter(col_idx)].width = 20
                    
                    # === Feuille 4: Analyse par Période Tarifaire ===
                    def classify_period(hour):
                        if hour in range(23, 24) or hour in range(0, 6):
                            return 'Heures Creuses'
                        elif hour in range(18, 23):
                            return 'Heures Pleines'
                        else:
                            return 'Heures Normales'
                    
                    df['period'] = df['hour'].apply(classify_period)
                    
                    period_analysis = df.groupby('period').agg({
                        'energy_consumption_kwh': ['sum', 'mean'],
                        'energy_cost_fcfa': ['sum', 'mean'],
                        'pump_efficiency': 'mean'
                    }).round(2)
                    
                    period_analysis.columns = ['_'.join(col).strip() for col in period_analysis.columns]
                    period_analysis.reset_index(inplace=True)
                    period_analysis.to_excel(writer, sheet_name='Analyse Tarifaire', index=False, startrow=2)
                    
                    ws_period = writer.sheets['Analyse Tarifaire']
                    self.format_excel_worksheet(ws_period, f"Station {station_name} - Analyse Tarifaire")
                    
                    for column in ws_period.columns:
                        ws_period.column_dimensions[get_column_letter(col_idx)].width = 20
                    
                    # === Feuille 5: Anomalies ===
                    if 'anomaly' in df.columns:
                        df_anomalies = df[df['anomaly'] == 1].copy()
                        if not df_anomalies.empty:
                            df_anomalies_export = df_anomalies[[
                                'timestamp', 'energy_consumption_kwh', 'pump_efficiency',
                                'power_factor', 'specific_consumption_kwh_m3'
                            ]].copy()
                            df_anomalies_export['timestamp'] = pd.to_datetime(
                                df_anomalies_export['timestamp']
                            ).dt.strftime('%Y-%m-%d %H:%M:%S')
                            
                            df_anomalies_export.to_excel(
                                writer, sheet_name='Anomalies Détectées', index=False, startrow=2
                            )
                            
                            ws_anomalies = writer.sheets['Anomalies Détectées']
                            self.format_excel_worksheet(
                                ws_anomalies,
                                f"Station {station_name} - Anomalies Détectées"
                            )
                            
                            for column in ws_anomalies.columns:
                                ws_anomalies.column_dimensions[get_column_letter(col_idx)].width = 22
                    
                    # === Feuille 6: Recommandations ===
                    recommendations = []
                    
                    # Analyser efficacité
                    avg_efficiency = df['pump_efficiency'].mean()
                    if avg_efficiency < 0.75:
                        recommendations.append({
                            'Catégorie': 'Performance',
                            'Priorité': 'Haute',
                            'Recommandation': 'Maintenance préventive urgente - efficacité < 75%',
                            'Impact Estimé': 'Amélioration efficacité +10-15%'
                        })
                    elif avg_efficiency < 0.80:
                        recommendations.append({
                            'Catégorie': 'Performance',
                            'Priorité': 'Moyenne',
                            'Recommandation': 'Surveiller efficacité des pompes',
                            'Impact Estimé': 'Prévention dégradation'
                        })
                    
                    # Analyser facteur de puissance
                    if 'power_factor' in df.columns:
                        avg_pf = df['power_factor'].mean()
                        if avg_pf < 0.85:
                            recommendations.append({
                                'Catégorie': 'Électrique',
                                'Priorité': 'Haute',
                                'Recommandation': f'Installer condensateurs (PF={avg_pf:.2f})',
                                'Impact Estimé': 'Réduction pénalités 80-90%'
                            })
                    
                    # Analyser heures de pointe
                    peak_cost = df[df['period'] == 'Heures Pleines']['energy_cost_fcfa'].sum()
                    total_cost = df['energy_cost_fcfa'].sum()
                    if peak_cost / total_cost > 0.35:
                        recommendations.append({
                            'Catégorie': 'Optimisation',
                            'Priorité': 'Haute',
                            'Recommandation': 'Déplacer production vers heures creuses',
                            'Impact Estimé': f'Économies potentielles: {(peak_cost * 0.4):,.0f} FCFA/période'
                        })
                    
                    # Analyser anomalies
                    if 'anomaly' in df.columns:
                        anomaly_rate = (df['anomaly'].sum() / len(df)) * 100
                        if anomaly_rate > 5:
                            recommendations.append({
                                'Catégorie': 'Maintenance',
                                'Priorité': 'Haute',
                                'Recommandation': f'Investigation anomalies ({anomaly_rate:.1f}% des données)',
                                'Impact Estimé': 'Prévention pannes majeures'
                            })
                    
                    if recommendations:
                        df_reco = pd.DataFrame(recommendations)
                        df_reco.to_excel(writer, sheet_name='Recommandations', index=False, startrow=2)
                        
                        ws_reco = writer.sheets['Recommandations']
                        self.format_excel_worksheet(
                            ws_reco,
                            f"Station {station_name} - Recommandations"
                        )
                        
                        ws_reco.column_dimensions['A'].width = 15
                        ws_reco.column_dimensions['B'].width = 12
                        ws_reco.column_dimensions['C'].width = 50
                        ws_reco.column_dimensions['D'].width = 35
                        
                        # Colorer les priorités
                        priority_colors = {
                            'Haute': 'FFB6B6',
                            'Moyenne': 'FFE6B6',
                            'Basse': 'B6FFB6'
                        }
                        
                        for row in range(4, 4 + len(recommendations)):
                            priority_cell = ws_reco[f'B{row}']
                            if priority_cell.value in priority_colors:
                                priority_cell.fill = PatternFill(
                                    start_color=priority_colors[priority_cell.value],
                                    end_color=priority_colors[priority_cell.value],
                                    fill_type="solid"
                                )
            
            logger.info(f"Export Excel créé: {filepath}")
            return str(filepath)
        
        except Exception as e:
            logger.error(f"Erreur export Excel: {str(e)}")
            raise
    
    def export_all_stations(
        self,
        station_ids: List[str],
        days: Optional[int] = 7
    ) -> str:
        """
        Exporter les données de toutes les stations dans un seul fichier Excel
        
        Args:
            station_ids: Liste des IDs de stations
            days: Nombre de jours de données
            
        Returns:
            Chemin du fichier Excel généré
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ONEA_Toutes_Stations_{timestamp}.xlsx"
            filepath = EXPORT_PATH / filename
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Feuille de synthèse
                summary_data = []
                
                for station_id in station_ids:
                    try:
                        df = self.load_station_data(station_id, days)
                        
                        summary_data.append({
                            'Station ID': station_id,
                            'Période': f"{df['timestamp'].min().strftime('%Y-%m-%d')} à {df['timestamp'].max().strftime('%Y-%m-%d')}",
                            'Points de données': len(df),
                            'Consommation totale (kWh)': df['energy_consumption_kwh'].sum(),
                            'Coût total (FCFA)': df['energy_cost_fcfa'].sum(),
                            'Efficacité moyenne (%)': df['pump_efficiency'].mean() * 100,
                            'Anomalies': df['anomaly'].sum() if 'anomaly' in df.columns else 0
                        })
                        
                        # Ajouter les données de chaque station dans une feuille séparée
                        df_export = df.copy()
                        df_export['timestamp'] = df_export['timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
                        df_export.to_excel(
                            writer,
                            sheet_name=station_id[:31],  # Limite Excel: 31 caractères
                            index=False
                        )
                    
                    except Exception as e:
                        logger.warning(f"Impossible d'exporter {station_id}: {e}")
                
                # Créer la feuille de synthèse
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Synthèse', index=False, startrow=2)
                
                ws_summary = writer.sheets['Synthèse']
                self.format_excel_worksheet(ws_summary, "ONEA - Synthèse Toutes Stations")
                
                for column in ws_summary.columns:
                    ws_summary.column_dimensions[get_column_letter(col_idx)].width = 22
            
            logger.info(f"Export multi-stations créé: {filepath}")
            return str(filepath)
        
        except Exception as e:
            logger.error(f"Erreur export multi-stations: {str(e)}")
            raise


# Instance globale du service
excel_export_service = ExcelExportService()
