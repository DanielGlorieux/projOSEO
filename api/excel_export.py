"""
Service d'export de données vers Excel
Pour téléchargement et analyse externe
"""
import pandas as pd
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, BarChart

logger = logging.getLogger(__name__)

DATA_PATH = Path(__file__).parent.parent / "data" / "raw"
EXPORT_PATH = Path(__file__).parent.parent / "data" / "exports"


class ExcelExportService:
    """Service d'export de datasets au format Excel"""
    
    def __init__(self):
        # Créer le dossier d'exports s'il n'existe pas
        EXPORT_PATH.mkdir(parents=True, exist_ok=True)
    
    def load_station_data(self, station_id: str, days: Optional[int] = None) -> pd.DataFrame:
        """Charger les données d'une station"""
        csv_path = DATA_PATH / f"{station_id}_historical.csv"
        
        if not csv_path.exists():
            raise FileNotFoundError(f"Données non trouvées pour {station_id}")
        
        df = pd.read_csv(csv_path)
        
        # Convertir timestamp et gérer les NaT/erreurs
        df['timestamp'] = pd.to_datetime(df['timestamp'], errors='coerce')
        
        # Supprimer les lignes avec timestamp invalide
        df = df.dropna(subset=['timestamp'])
        
        if len(df) == 0:
            raise ValueError(f"Aucune donnée valide pour {station_id}")
        
        df = df.sort_values('timestamp')
        
        # Filtrer par période si spécifié
        if days:
            # Utiliser la date la plus récente des données comme référence
            max_date = df['timestamp'].max()
            cutoff_date = max_date - timedelta(days=days)
            df = df[df['timestamp'] >= cutoff_date]
            logger.info(f"Filtrage {days} jours depuis {max_date}: {len(df)} lignes")
        
        return df
    
    def format_excel_worksheet(self, ws, title: str = None, header_row: int = 1):
        """Formatter une feuille Excel avec styles professionnels (en-têtes seulement)"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell import MergedCell
        
        # En-têtes de colonnes
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formater la ligne d'en-tête en évitant les cellules fusionnées
        try:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                
                # Ignorer les cellules fusionnées
                if isinstance(cell, MergedCell):
                    continue
                    
                if cell.value:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
        except Exception as e:
            logger.warning(f"Erreur formatage en-têtes: {e}")
        
        # Ajuster la hauteur de la ligne d'en-tête
        try:
            ws.row_dimensions[header_row].height = 25
        except Exception as e:
            logger.warning(f"Erreur ajustement hauteur: {e}")
    
    def safe_strftime(self, timestamp) -> str:
        """Convertir timestamp en string de manière sécurisée"""
        try:
            if pd.isna(timestamp):
                return "N/A"
            return timestamp.strftime("%Y-%m-%d %H:%M:%S")
        except:
            return "N/A"
    
    def export_station_data(
        self,
        station_id: str,
        station_name: str,
        days: Optional[int] = 7,
        include_analytics: bool = True
    ) -> str:
        """
        Exporter les données d'une station au format Excel avec analyses
        
        Args:
            station_id: ID de la station
            station_name: Nom de la station
            days: Nombre de jours de données (None = toutes)
            include_analytics: Inclure feuilles d'analyse
            
        Returns:
            Chemin du fichier Excel généré
        """
        try:
            # Charger les données
            df = self.load_station_data(station_id, days)
            
            # Vérifier les colonnes nécessaires
            required_cols = ['timestamp', 'energy_consumption_kwh', 'energy_cost_fcfa', 'pump_efficiency']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                logger.warning(f"Colonnes manquantes: {missing_cols}, export basique seulement")
                include_analytics = False
            
            # Nom du fichier
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ONEA_{station_id}_{timestamp}.xlsx"
            filepath = EXPORT_PATH / filename
            
            # Créer le workbook Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # === Feuille 1: Données brutes ===
                df_export = df.copy()
                
                # Convertir timestamp de manière sûre (gérer NaT)
                df_export['timestamp'] = df_export['timestamp'].apply(
                    lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and hasattr(x, 'strftime') else ''
                )
                
                # Écrire les données
                df_export.to_excel(writer, sheet_name='Données Brutes', index=False)
                
                # Formater la feuille
                ws_raw = writer.sheets['Données Brutes']
                self.format_excel_worksheet(ws_raw, header_row=1)
                
                # Ajuster largeur des colonnes
                try:
                    from openpyxl.utils import get_column_letter
                    for col_idx in range(1, ws_raw.max_column + 1):
                        column_letter = get_column_letter(col_idx)
                        ws_raw.column_dimensions[column_letter].width = 20
                except Exception as e:
                    logger.warning(f"Erreur ajustement colonnes: {e}")
                
                if include_analytics:
                    # === Feuille 2: Résumé Statistique ===
                    # Calculer les valeurs de manière sécurisée
                    try:
                        min_ts = df['timestamp'].min()
                        max_ts = df['timestamp'].max()
                        period_str = f"{min_ts:%Y-%m-%d} à {max_ts:%Y-%m-%d}" if pd.notna(min_ts) and pd.notna(max_ts) else "N/A"
                    except:
                        period_str = "N/A"
                    
                    stats_data = {
                        'Métrique': [
                            'Période analysée',
                            'Nombre de points de données',
                            'Consommation totale (kWh)',
                            'Consommation moyenne (kWh)',
                            'Coût total (FCFA)',
                            'Coût moyen (FCFA)',
                            'Efficacité moyenne (%)',
                            'Nombre d\'anomalies',
                            'Pompes actives (moyenne)',
                            'Facteur de puissance moyen'
                        ],
                        'Valeur': [
                            period_str,
                            str(len(df)),
                            f"{df['energy_consumption_kwh'].sum():,.2f}",
                            f"{df['energy_consumption_kwh'].mean():,.2f}",
                            f"{df['energy_cost_fcfa'].sum():,.2f}",
                            f"{df['energy_cost_fcfa'].mean():,.2f}",
                            f"{df['pump_efficiency'].mean() * 100:.2f}",
                            str(int(df['anomaly'].sum()) if 'anomaly' in df.columns else 0),
                            f"{df['pumps_active'].mean():.2f}",
                            f"{df['power_factor'].mean():.3f}" if 'power_factor' in df.columns else "N/A"
                        ]
                    }
                    
                    df_stats = pd.DataFrame(stats_data)
                    df_stats.to_excel(writer, sheet_name='Résumé Statistique', index=False)
                    
                    ws_stats = writer.sheets['Résumé Statistique']
                    
                    # Ajouter le titre
                    ws_stats.insert_rows(1)
                    ws_stats['A1'] = f"Station {station_name} - Résumé Statistique"
                    ws_stats['A1'].font = Font(bold=True, size=14, color="1F4788")
                    ws_stats.merge_cells('A1:B1')
                    
                    self.format_excel_worksheet(ws_stats, header_row=2)
                    ws_stats.column_dimensions['A'].width = 35
                    ws_stats.column_dimensions['B'].width = 30
                    
                    # === Feuille 3: Analyse par Heure ===
                    df['hour'] = pd.to_datetime(df['timestamp']).dt.hour
                    hourly_analysis = df.groupby('hour').agg({
                        'energy_consumption_kwh': ['mean', 'min', 'max', 'std'],
                        'energy_cost_fcfa': ['mean', 'sum'],
                        'pump_efficiency': 'mean',
                        'pumps_active': 'mean'
                    }).round(2)
                    
                    hourly_analysis.columns = ['_'.join(col).strip() for col in hourly_analysis.columns]
                    hourly_analysis.reset_index(inplace=True)
                    hourly_analysis.to_excel(writer, sheet_name='Analyse Horaire', index=False)
                    
                    ws_hourly = writer.sheets['Analyse Horaire']
                    
                    # Ajouter le titre
                    ws_hourly.insert_rows(1)
                    ws_hourly['A1'] = f"Station {station_name} - Analyse par Heure"
                    ws_hourly['A1'].font = Font(bold=True, size=14, color="1F4788")
                    
                    self.format_excel_worksheet(ws_hourly, header_row=2)
                    
                    # Ajuster largeur colonnes
                    try:
                        from openpyxl.utils import get_column_letter
                        for col_idx in range(1, ws_hourly.max_column + 1):
                            column_letter = get_column_letter(col_idx)
                            ws_hourly.column_dimensions[column_letter].width = 20
                    except Exception as e:
                        logger.warning(f"Erreur ajustement colonnes hourly: {e}")
                    
                    # === Feuille 4: Analyse par Période Tarifaire ===
                    def classify_period(hour):
                        if hour in range(23, 24) or hour in range(0, 6):
                            return 'Heures Creuses'
                        elif hour in range(18, 23):
                            return 'Heures Pleines'
                        else:
                            return 'Heures Normales'
                    
                    df['period'] = df['hour'].apply(classify_period)
                    
                    period_analysis = df.groupby('period').agg({
                        'energy_consumption_kwh': ['sum', 'mean'],
                        'energy_cost_fcfa': ['sum', 'mean'],
                        'pump_efficiency': 'mean'
                    }).round(2)
                    
                    period_analysis.columns = ['_'.join(col).strip() for col in period_analysis.columns]
                    period_analysis.reset_index(inplace=True)
                    period_analysis.to_excel(writer, sheet_name='Analyse Tarifaire', index=False)
                    
                    ws_period = writer.sheets['Analyse Tarifaire']
                    
                    # Ajouter le titre
                    ws_period.insert_rows(1)
                    ws_period['A1'] = f"Station {station_name} - Analyse Tarifaire"
                    ws_period['A1'].font = Font(bold=True, size=14, color="1F4788")
                    
                    self.format_excel_worksheet(ws_period, header_row=2)
                    
                    # Ajuster largeur colonnes
                    try:
                        from openpyxl.utils import get_column_letter
                        for col_idx in range(1, ws_period.max_column + 1):
                            column_letter = get_column_letter(col_idx)
                            ws_period.column_dimensions[column_letter].width = 20
                    except Exception as e:
                        logger.warning(f"Erreur ajustement colonnes period: {e}")
                    
                    # === Feuille 5: Anomalies ===
                    if 'anomaly' in df.columns:
                        df_anomalies = df[df['anomaly'] == 1].copy()
                        if not df_anomalies.empty:
                            df_anomalies_export = df_anomalies[[
                                'timestamp', 'energy_consumption_kwh', 'pump_efficiency',
                                'power_factor', 'specific_consumption_kwh_m3'
                            ]].copy()
                            df_anomalies_export['timestamp'] = df_anomalies_export['timestamp'].apply(
                                lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) and hasattr(x, 'strftime') else ''
                            )
                            
                            df_anomalies_export.to_excel(
                                writer, sheet_name='Anomalies Détectées', index=False
                            )
                            
                            ws_anomalies = writer.sheets['Anomalies Détectées']
                            
                            # Ajouter le titre
                            ws_anomalies.insert_rows(1)
                            ws_anomalies['A1'] = f"Station {station_name} - Anomalies Détectées"
                            ws_anomalies['A1'].font = Font(bold=True, size=14, color="DC143C")
                            
                            self.format_excel_worksheet(ws_anomalies, header_row=2)
                            
                        # Ajuster largeur colonnes
                        try:
                            from openpyxl.utils import get_column_letter
                            for col_idx in range(1, ws_anomalies.max_column + 1):
                                column_letter = get_column_letter(col_idx)
                                ws_anomalies.column_dimensions[column_letter].width = 22
                        except Exception as e:
                            logger.warning(f"Erreur ajustement colonnes anomalies: {e}")
                    
                    # === Feuille 6: Recommandations ===
                    recommendations = []
                    
                    # Analyser efficacité
                    avg_efficiency = df['pump_efficiency'].mean()
                    if avg_efficiency < 0.75:
                        recommendations.append({
                            'Catégorie': 'Performance',
                            'Priorité': 'Haute',
                            'Recommandation': 'Maintenance préventive urgente - efficacité < 75%',
                            'Impact Estimé': 'Amélioration efficacité +10-15%'
                        })
                    elif avg_efficiency < 0.80:
                        recommendations.append({
                            'Catégorie': 'Performance',
                            'Priorité': 'Moyenne',
                            'Recommandation': 'Surveiller efficacité des pompes',
                            'Impact Estimé': 'Prévention dégradation'
                        })
                    
                    # Analyser facteur de puissance
                    if 'power_factor' in df.columns:
                        avg_pf = df['power_factor'].mean()
                        if avg_pf < 0.85:
                            recommendations.append({
                                'Catégorie': 'Électrique',
                                'Priorité': 'Haute',
                                'Recommandation': f'Installer condensateurs (PF={avg_pf:.2f})',
                                'Impact Estimé': 'Réduction pénalités 80-90%'
                            })
                    
                    # Analyser heures de pointe
                    peak_cost = df[df['period'] == 'Heures Pleines']['energy_cost_fcfa'].sum()
                    total_cost = df['energy_cost_fcfa'].sum()
                    if peak_cost / total_cost > 0.35:
                        recommendations.append({
                            'Catégorie': 'Optimisation',
                            'Priorité': 'Haute',
                            'Recommandation': 'Déplacer production vers heures creuses',
                            'Impact Estimé': f'Économies potentielles: {(peak_cost * 0.4):,.0f} FCFA/période'
                        })
                    
                    # Analyser anomalies
                    if 'anomaly' in df.columns:
                        anomaly_rate = (df['anomaly'].sum() / len(df)) * 100
                        if anomaly_rate > 5:
                            recommendations.append({
                                'Catégorie': 'Maintenance',
                                'Priorité': 'Haute',
                                'Recommandation': f'Investigation anomalies ({anomaly_rate:.1f}% des données)',
                                'Impact Estimé': 'Prévention pannes majeures'
                            })
                    
                    if recommendations:
                        df_reco = pd.DataFrame(recommendations)
                        df_reco.to_excel(writer, sheet_name='Recommandations', index=False)
                        
                        ws_reco = writer.sheets['Recommandations']
                        
                        # Ajouter le titre
                        ws_reco.insert_rows(1)
                        ws_reco['A1'] = f"Station {station_name} - Recommandations"
                        ws_reco['A1'].font = Font(bold=True, size=14, color="1F4788")
                        
                        self.format_excel_worksheet(ws_reco, header_row=2)
                        
                        ws_reco.column_dimensions['A'].width = 15
                        ws_reco.column_dimensions['B'].width = 12
                        ws_reco.column_dimensions['C'].width = 50
                        ws_reco.column_dimensions['D'].width = 35
                        
                        # Colorer les priorités
                        priority_colors = {
                            'Haute': 'FFB6B6',
                            'Moyenne': 'FFE6B6',
                            'Basse': 'B6FFB6'
                        }
                        
                        for row in range(3, 3 + len(recommendations)):  # Commencer à row 3 car titre en row 1
                            priority_cell = ws_reco[f'B{row}']
                            if priority_cell.value in priority_colors:
                                priority_cell.fill = PatternFill(
                                    start_color=priority_colors[priority_cell.value],
                                    end_color=priority_colors[priority_cell.value],
                                    fill_type="solid"
                                )
            
            logger.info(f"Export Excel créé: {filepath}")
            return str(filepath)
        
        except Exception as e:
            logger.error(f"Erreur export Excel: {str(e)}")
            raise
    
    def export_all_stations(
        self,
        station_ids: List[str],
        days: Optional[int] = 7
    ) -> str:
        """
        Exporter les données de toutes les stations dans un seul fichier Excel
        
        Args:
            station_ids: Liste des IDs de stations
            days: Nombre de jours de données
            
        Returns:
            Chemin du fichier Excel généré
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ONEA_Toutes_Stations_{timestamp}.xlsx"
            filepath = EXPORT_PATH / filename
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Feuille de synthèse
                summary_data = []
                
                for station_id in station_ids:
                    try:
                        df = self.load_station_data(station_id, days)
                        
                        # Convertir les timestamps de manière sûre
                        min_ts = df['timestamp'].min()
                        max_ts = df['timestamp'].max()
                        
                        if pd.notna(min_ts) and pd.notna(max_ts):
                            period_str = f"{min_ts:%Y-%m-%d} à {max_ts:%Y-%m-%d}"
                        else:
                            period_str = "N/A"
                        
                        summary_data.append({
                            'Station ID': station_id,
                            'Période': period_str,
                            'Points de données': len(df),
                            'Consommation totale (kWh)': df['energy_consumption_kwh'].sum(),
                            'Coût total (FCFA)': df['energy_cost_fcfa'].sum(),
                            'Efficacité moyenne (%)': df['pump_efficiency'].mean() * 100,
                            'Anomalies': df['anomaly'].sum() if 'anomaly' in df.columns else 0
                        })
                        
                        # Ajouter les données de chaque station dans une feuille séparée
                        df_export = df.copy()
                        df_export['timestamp'] = df_export['timestamp'].apply(
                            lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(x) else ''
                        )
                        df_export.to_excel(
                            writer,
                            sheet_name=station_id[:31],  # Limite Excel: 31 caractères
                            index=False
                        )
                    
                    except Exception as e:
                        logger.warning(f"Impossible d'exporter {station_id}: {e}")
                
                # Créer la feuille de synthèse
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Synthèse', index=False)
                
                ws_summary = writer.sheets['Synthèse']
                
                # Ajouter le titre
                ws_summary.insert_rows(1)
                ws_summary['A1'] = "ONEA - Synthèse Toutes Stations"
                ws_summary['A1'].font = Font(bold=True, size=14, color="1F4788")
                
                self.format_excel_worksheet(ws_summary, header_row=2)
                
                # Ajuster largeur colonnes
                try:
                    from openpyxl.utils import get_column_letter
                    for col_idx in range(1, ws_summary.max_column + 1):
                        column_letter = get_column_letter(col_idx)
                        ws_summary.column_dimensions[column_letter].width = 22
                except Exception as e:
                    logger.warning(f"Erreur ajustement colonnes summary: {e}")
            
            logger.info(f"Export multi-stations créé: {filepath}")
            return str(filepath)
        
        except Exception as e:
            logger.error(f"Erreur export multi-stations: {str(e)}")
            raise


# Instance globale du service
excel_export_service = ExcelExportService()
